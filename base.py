# -*- coding: utf-8 -*-
import os
from copy import copy

import openpyxl
import time
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from time import sleep

from constants import (
    NAME_FILE, NAME_PAGE, MAP_EXCEL, DATA_ROW_HEADER, FIRST, CP_1251,
    MAP_MONTH,
    PHONE, COL_NAMES, RANGE_WITH_SUM, START, TEMPLATE_RANGE, END,
    AMOUNT_PER_NUMBER, ONE, CONTRACT, PATH_TO_IN, XLSM, ZERO)
from logger import ReportLoggerMixin
from report import PDFReport, BillingErrorFileWrite


def print_progress_bar(
        iteration, total, prefix='', suffix='',
        decimals=1, length=100, fill='#', time_sec=0
):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration:  Номер текущей итерации
        total:      Количество итераций
        prefix:     Сообщение в строке прогресса , перед прогресс баром
        suffix:     Сообщение в строке прогресса , после прогресс бара
        decimals:   Целое, позитовное число прогесса
        length:     Длина в символах , прогесс бара
        fill:       символ заполнения прогресс бара
        time_sec:   Время в секундах
    """
    percent = (
            "{0:." + str(decimals) + "f}"
    ).format(100 * (iteration / float(total)))
    filled_length = int(length * iteration // total)
    progress_time = time.time() - time_sec
    bar = fill * filled_length + '-' * (length - filled_length)
    sys.stdout.write('\r%s |%s| %s%% %s\t\t%f sec' % (
        prefix, bar, percent, suffix, progress_time))
    if iteration == total:
        sys.stdout.write('\n')


class AllReports(PDFReport):
    END_LINE = u'Итого по лицевому счету без скидки'
    DICT_ARGUMENTS = {
        # Номер группы : Имя группы
        # Номер группы в регулярном вырожении
        12: AMOUNT_PER_NUMBER,
        2: PHONE
    }

    def __init__(self, data_unit, pattern_date, pattern_contract, out_xls):
        super(
            AllReports,
            self).__init__(data_unit, pattern_date, pattern_contract)
        self.out_file = out_xls

    def write_(self, out_file, result):
        book_excel = openpyxl.load_workbook(out_file)

        if XLSM in out_file:
            book_excel = openpyxl.load_workbook(out_file, keep_vba=True)

        page_for_report = book_excel[NAME_PAGE]
        day, month, year = self.date_.split('.')
        month_name = MAP_MONTH.get(month)
        range_with_data_headers = page_for_report[MAP_EXCEL.get(
            DATA_ROW_HEADER)
        ]
        cell_phone = MAP_EXCEL[COL_NAMES][PHONE]
        cell_phone_start = MAP_EXCEL[RANGE_WITH_SUM][START]
        cell_phone_end = MAP_EXCEL[RANGE_WITH_SUM][END]

        phone_range = TEMPLATE_RANGE.format(
            cell_char_one=cell_phone,
            digit_one=cell_phone_start,
            cell_char_two=cell_phone,
            digit_two=cell_phone_end
        )
        range_phone_all = page_for_report[phone_range]

        range_date_headers = filter(
            lambda x:
            x if x.value == ' '.join([month_name, year]) else None,
            range_with_data_headers[FIRST]
        )

        self.write_log(ReportLoggerMixin.INFO,
                       u"Успех: Выборка дат из excel завершина")

        range_phone = filter(lambda x: x[0].value is not None, range_phone_all)

        self.write_log(ReportLoggerMixin.INFO,
                       u"Успех: Выборка телефонов завершина.")

        if range_date_headers:
            cell_cur = range_date_headers[ZERO]

            self.write_log(
                ReportLoggerMixin.INFO,
                (
                    u"Успех: дата присуствует в"
                    u" списке ячейка {}{} значение ".format(
                        cell_cur.column, cell_cur.row,
                        cell_cur.value)
                )
            )

            phones_amount_dict = map(
                lambda x: {
                    int(x[PHONE]):
                        round(
                            float(
                                x[AMOUNT_PER_NUMBER].replace(',', '.')
                            ),
                            2
                        )
                },
                result)

            dict_res = {}
            amount_per_number = None
            for item in phones_amount_dict:
                dict_res.update(item)
            for item_cell in range_phone:
                var_temp_value = int(item_cell[0].value)
                var_temp_row_index = '{}{}'.format(
                    cell_cur.column, item_cell[0].row)
                try:
                    cur_value = float(
                        page_for_report[var_temp_row_index].value or 0.0
                    )
                except ValueError:
                    cur_value = 0.0

                amount_per_number = dict_res.get(var_temp_value, None)
                if amount_per_number is not None:
                    dict_res.pop(var_temp_value)
                    self.write_log(
                        ReportLoggerMixin.INFO,
                        u"Информация: "
                        u"Подстановка значений {}+"
                        u"{} в ячейку {}".format(
                            amount_per_number, cur_value,
                            var_temp_row_index)
                    )
                    page_for_report[var_temp_row_index].value = (
                            amount_per_number + cur_value)

                    left_cell = chr(ord(cell_cur.column)-1)
                    fill_ = page_for_report[
                        '{}{}'.format(left_cell, item_cell[0].row)].fill

                    color = fill_.fgColor.rgb
                    if color == '00000000':
                        # Замена черного на белый
                        color = 'FFFFFFFF'
                    try:
                        left_fill = PatternFill(
                            fill_type='solid',
                            fgColor=Color(color),
                        )
                    except TypeError:
                        # Не всегда удается определить цвет

                        new_cell = page_for_report[var_temp_row_index]
                        old_cell = page_for_report[
                            '{}{}'.format(left_cell, item_cell[ZERO].row)]
                        # копируем значения стилей
                        new_cell.font = copy(old_cell.font)
                        new_cell.border = copy(old_cell.border)
                        new_cell.fill = copy(old_cell.fill)
                        new_cell.number_format = copy(old_cell.number_format)
                        new_cell.protection = copy(old_cell.protection)
                        new_cell.alignment = copy(old_cell.alignment)

                    else:
                        page_for_report[var_temp_row_index].fill = left_fill

            last_num = range_phone[-1][0].row
            if dict_res and not amount_per_number:
                for item_not_found in dict_res:

                    self.write_log(
                        ReportLoggerMixin.INFO,
                        u"Информация: "
                        u"{}   нет в списке".format(item_not_found)
                    )

                    last_num += ONE

                    self.write_log(
                        ReportLoggerMixin.INFO,
                        u"Информация: "
                        u"добавляю значение {} в {}{}".format(
                            dict_res[item_not_found],
                            cell_cur.column, last_num
                        )
                    )

                    page_for_report[
                        '{}{}'.format(cell_cur.column, last_num)
                    ].value = dict_res[item_not_found]
                    self.write_log(
                        ReportLoggerMixin.INFO,
                        u"Информация"
                        u"добавляю значение {} в {}{}".format(
                            item_not_found, MAP_EXCEL[COL_NAMES][PHONE],
                            last_num
                       )
                    )

                    page_for_report[
                        '{}{}'.format(MAP_EXCEL[COL_NAMES][PHONE], last_num)
                    ].value = item_not_found

                    self.write_log(
                        ReportLoggerMixin.INFO,
                        u"Информация: добавляю значение {} в {}{}".format(
                            self.contract, MAP_EXCEL[COL_NAMES][CONTRACT],
                            last_num)
                    )
                    page_for_report[
                        '{}{}'.format(
                            MAP_EXCEL[COL_NAMES][CONTRACT], last_num)
                    ].value = self.contract

        else:

            self.write_log(
                ReportLoggerMixin.WARNING,
                u"Внимание: дата {} {} {} не найдена ...".format(
                    day, month, year)
            )
        try:
            book_excel.save(out_file)
            book_excel.close()

        except IOError:
            self.write_log(
                ReportLoggerMixin.ERROR,
                u"Ошибка: Файл используется другой программой"
            )
            raise BillingErrorFileWrite


if __name__ == "__main__" and __package__ is None:
    from os import sys, path

    try:
        sys.path.append(path.dirname(path.dirname(path.abspath(__file__))))
    except NameError:  # py2exe script, not a module
        import sys
        sys.path.append(path.dirname(path.dirname(sys.argv[0])))

    files = os.listdir(PATH_TO_IN)
    sys.stdout.write(u"Авто заполнение отчета... \n")
    sys.stdout.flush()
    iter_progress = ZERO
    now = time.time()
    for file_ in files:
        print_progress_bar(iter_progress + 1, len(files),
                           prefix='Progress:', suffix='Complete',
                           length=50, time_sec=now)
        iter_progress += 1
        try:
            all_reports = AllReports(
                (
                    u'(Сетевой ресурс(\d{11})'
                    u'\s(Периодические\sуслуги\s\d{1,6},'
                    u'\d{4}\s{2})?(Скидка\soff-line'
                    u'\s\(Периодические услуги\)-\d{1,6},'
                    u'\d{4}\s{2})?(Разовые\sуслуги'
                    u'\s\d{1,6},\d{4})?(Скидка\soff-line\s'
                    u'\(Разовые\sуслуги\)-\d{1,6},\d{4}\s{2})'
                    u'?(Телефонные\sуслуги\s\d{1,6},'
                    u'\d{4}\s{0,2})?(Скидка\soff-line\s\(Телефонные\sуслуги\)-'
                    u'\d{1,6},\d{4}\s{2})?(НДС\s\d{1,6},\d{4})'
                    u'?(Итого\s\без\sскидки:'
                    u'\s\d{1,6},\d{4}\s{2})'
                    u'?(Итого\sскидка\soff-line:-\d{1,6},\d{4}\s{2})?'
                    u'Итого:\s(\d{1,6},\d{4})?)'
                ),

                u'.+(?P<date_>\d{2}\.\d{2}\.\d{4})',

                u'.+\d{12}\s{2}(?P<contract>\d{3}\s\d{3}\s\d{3}\s\d{3})',
                NAME_FILE
            )
            all_reports.find_values(os.path.join(PATH_TO_IN, file_))

            all_reports.write_log(
                ReportLoggerMixin.INFO,
                u'Работа с Файлом : {} завершина\n\n\n\n'.format(
                    file_.decode(CP_1251)
                )
            )
        except Exception as err:
            sys.stdout.write(err.message)

    sys.stdout.write(u" Завершино ... \n")
    all_reports.write_log(ReportLoggerMixin.INFO, u'Завершино ...')
    sys.stdout.write(u" Для выхода зажмите ctrl + C ")
    try:
        sleep(60)
    except KeyboardInterrupt:
        sys.exit(0)
